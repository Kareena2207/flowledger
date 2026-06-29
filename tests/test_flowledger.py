from __future__ import annotations

import tempfile
import unittest
from io import StringIO
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from flowledger.csv_io import parse_invoices, parse_payments
from flowledger.engine import build_report_data
from flowledger.models import Invoice, Payment
from flowledger.workbook import create_workbook, create_workbook_bytes


class FlowLedgerTests(unittest.TestCase):
    def setUp(self) -> None:
        self.invoices = [
            Invoice("INV-001", "Northwind Studio", date(2026, 6, 1), date(2026, 6, 10), Decimal("1200.00"), "USD"),
            Invoice("INV-002", "Maple Labs", date(2026, 6, 12), date(2026, 6, 28), Decimal("800.00"), "USD"),
            Invoice("INV-003", "Northwind Studio", date(2026, 6, 15), date(2026, 7, 4), Decimal("500.00"), "USD"),
        ]
        self.payments = [
            Payment("PAY-001", "INV-001", date(2026, 6, 9), Decimal("1200.00"), "Bank Transfer", ""),
            Payment("PAY-002", "INV-002", date(2026, 6, 20), Decimal("300.00"), "UPI", "Partial payment"),
        ]

    def test_build_report_data_derives_status_and_metrics(self) -> None:
        report = build_report_data(self.invoices, self.payments, date(2026, 6, 29))

        self.assertEqual(report["dashboard"]["invoice_count"], 3)
        self.assertEqual(report["dashboard"]["pending_total"], Decimal("1000.00"))
        self.assertEqual(report["dashboard"]["overdue_total"], Decimal("500.00"))

        invoice_rows = {row["invoice_id"]: row for row in report["invoices"]}
        self.assertEqual(invoice_rows["INV-001"]["status"], "paid")
        self.assertEqual(invoice_rows["INV-002"]["status"], "overdue")
        self.assertEqual(invoice_rows["INV-003"]["status"], "due_soon")
        self.assertEqual(invoice_rows["INV-002"]["outstanding_amount"], Decimal("500.00"))

    def test_create_workbook_writes_expected_sheets(self) -> None:
        report = build_report_data(self.invoices, self.payments, date(2026, 6, 29))
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "flowledger.xlsx"
            create_workbook(report, output)

            workbook = load_workbook(output)
            self.assertEqual(
                workbook.sheetnames,
                ["Dashboard", "Invoices", "Payments", "Clients", "Monthly Summary", "Alerts"],
            )
            self.assertEqual(workbook["Dashboard"]["A1"].value, "FlowLedger Dashboard")
            self.assertEqual(workbook["Invoices"]["A4"].value, "INV-001")

    def test_create_workbook_bytes_returns_xlsx_content(self) -> None:
        report = build_report_data(self.invoices, self.payments, date(2026, 6, 29))
        workbook_bytes = create_workbook_bytes(report)

        self.assertTrue(workbook_bytes.startswith(b"PK"))

    def test_parse_uploaded_csv_text(self) -> None:
        invoices_csv = StringIO(
            "invoice_id,client_name,issue_date,due_date,amount,currency,status\n"
            "INV-100,Acme,2026-06-01,2026-06-15,1000.00,USD,\n"
        )
        payments_csv = StringIO(
            "payment_id,invoice_id,payment_date,amount,method,notes\n"
            "PAY-100,INV-100,2026-06-10,1000.00,ACH,Full payment\n"
        )

        invoices = parse_invoices(invoices_csv)
        payments = parse_payments(payments_csv)

        self.assertEqual(invoices[0].invoice_id, "INV-100")
        self.assertEqual(payments[0].payment_id, "PAY-100")


if __name__ == "__main__":
    unittest.main()
