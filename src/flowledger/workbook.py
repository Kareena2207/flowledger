from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OVERDUE_FILL = PatternFill("solid", fgColor="FDE9E7")
DUE_SOON_FILL = PatternFill("solid", fgColor="FFF2CC")
PAID_FILL = PatternFill("solid", fgColor="E2F0D9")


def _auto_fit(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(length + 2, 12), 28
        )


def _write_table(ws, title: str, headers: list[str], rows: list[list[object]]) -> None:
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A4"
    _auto_fit(ws)


def _build_workbook(report_data: dict[str, object]) -> Workbook:
    workbook = Workbook()
    dashboard_ws = workbook.active
    dashboard_ws.title = "Dashboard"

    dashboard = report_data["dashboard"]
    metrics = [
        ("As Of", dashboard["as_of"]),
        ("Invoices", dashboard["invoice_count"]),
        ("Clients", dashboard["client_count"]),
        ("Collected This Month", dashboard["collected_this_month"]),
        ("Pending Total", dashboard["pending_total"]),
        ("Overdue Total", dashboard["overdue_total"]),
        ("Forecast 30 Day Inflow", dashboard["forecast_30_day_inflow"]),
    ]

    dashboard_ws["A1"] = "FlowLedger Dashboard"
    dashboard_ws["A1"].font = Font(bold=True, size=16)
    for index, (label, value) in enumerate(metrics, start=3):
        dashboard_ws[f"A{index}"] = label
        dashboard_ws[f"B{index}"] = value
        if isinstance(value, Decimal):
            dashboard_ws[f"B{index}"].number_format = "$#,##0.00"
    dashboard_ws.column_dimensions["A"].width = 24
    dashboard_ws.column_dimensions["B"].width = 18

    monthly_rows = report_data["monthly_summary"]
    dashboard_ws["D2"] = "Month"
    dashboard_ws["E2"] = "Invoiced"
    dashboard_ws["F2"] = "Collected"
    for cell in dashboard_ws[2][3:6]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_index, row in enumerate(monthly_rows, start=3):
        dashboard_ws[f"D{row_index}"] = row["month"]
        dashboard_ws[f"E{row_index}"] = row["invoiced_amount"]
        dashboard_ws[f"F{row_index}"] = row["collected_amount"]
        dashboard_ws[f"E{row_index}"].number_format = "$#,##0.00"
        dashboard_ws[f"F{row_index}"].number_format = "$#,##0.00"

    if monthly_rows:
        line_chart = LineChart()
        line_chart.title = "Monthly Cash Flow"
        line_chart.y_axis.title = "USD"
        line_chart.x_axis.title = "Month"
        data = Reference(dashboard_ws, min_col=5, max_col=6, min_row=2, max_row=2 + len(monthly_rows))
        categories = Reference(dashboard_ws, min_col=4, min_row=3, max_row=2 + len(monthly_rows))
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.height = 7
        line_chart.width = 12
        dashboard_ws.add_chart(line_chart, "D8")

    invoices_ws = workbook.create_sheet("Invoices")
    invoice_headers = [
        "Invoice ID",
        "Client",
        "Issue Date",
        "Due Date",
        "Amount",
        "Paid",
        "Outstanding",
        "Status",
        "Days Overdue",
    ]
    invoice_rows = [
        [
            row["invoice_id"],
            row["client_name"],
            row["issue_date"],
            row["due_date"],
            row["amount"],
            row["paid_amount"],
            row["outstanding_amount"],
            row["status"],
            row["days_overdue"],
        ]
        for row in report_data["invoices"]
    ]
    _write_table(invoices_ws, "Invoices", invoice_headers, invoice_rows)
    for row in invoices_ws.iter_rows(min_row=4, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "$#,##0.00"
    for row in invoices_ws.iter_rows(min_row=4, max_col=8):
        status_cell = row[7]
        if status_cell.value == "overdue":
            status_cell.fill = OVERDUE_FILL
        elif status_cell.value == "due_soon":
            status_cell.fill = DUE_SOON_FILL
        elif status_cell.value == "paid":
            status_cell.fill = PAID_FILL

    payments_ws = workbook.create_sheet("Payments")
    payment_headers = ["Payment ID", "Invoice ID", "Payment Date", "Amount", "Method", "Notes"]
    payment_rows = [
        [
            row["payment_id"],
            row["invoice_id"],
            row["payment_date"],
            row["amount"],
            row["method"],
            row["notes"],
        ]
        for row in report_data["payments"]
    ]
    _write_table(payments_ws, "Payments", payment_headers, payment_rows)
    for cell in payments_ws["D"][3:]:
        cell.number_format = "$#,##0.00"

    clients_ws = workbook.create_sheet("Clients")
    client_headers = ["Client", "Invoices", "Billed", "Collected", "Outstanding"]
    client_rows = [
        [
            row["client_name"],
            row["invoice_count"],
            row["billed_amount"],
            row["paid_amount"],
            row["outstanding_amount"],
        ]
        for row in report_data["clients"]
    ]
    _write_table(clients_ws, "Clients", client_headers, client_rows)
    for row in clients_ws.iter_rows(min_row=4, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "$#,##0.00"

    if report_data["clients"]:
        chart = BarChart()
        chart.title = "Top Client Billing"
        chart.y_axis.title = "USD"
        data = Reference(clients_ws, min_col=3, max_col=3, min_row=3, max_row=3 + len(client_rows))
        categories = Reference(clients_ws, min_col=1, min_row=4, max_row=3 + len(client_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 11
        clients_ws.add_chart(chart, "G4")

    monthly_ws = workbook.create_sheet("Monthly Summary")
    monthly_headers = ["Month", "Invoiced", "Collected"]
    monthly_sheet_rows = [
        [row["month"], row["invoiced_amount"], row["collected_amount"]]
        for row in report_data["monthly_summary"]
    ]
    _write_table(monthly_ws, "Monthly Summary", monthly_headers, monthly_sheet_rows)
    for row in monthly_ws.iter_rows(min_row=4, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "$#,##0.00"

    alerts_ws = workbook.create_sheet("Alerts")
    alert_headers = ["Type", "Invoice ID", "Client", "Message", "Amount"]
    alert_rows = [
        [row["type"], row["invoice_id"], row["client_name"], row["message"], row["amount"]]
        for row in report_data["alerts"]
    ]
    _write_table(alerts_ws, "Alerts", alert_headers, alert_rows)
    for cell in alerts_ws["E"][3:]:
        cell.number_format = "$#,##0.00"
    for row in alerts_ws.iter_rows(min_row=4, max_col=1):
        if row[0].value == "overdue":
            row[0].fill = OVERDUE_FILL
        elif row[0].value == "due_soon":
            row[0].fill = DUE_SOON_FILL

    return workbook


def create_workbook(report_data: dict[str, object], output_path: str | Path) -> Path:
    workbook = _build_workbook(report_data)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def create_workbook_bytes(report_data: dict[str, object]) -> bytes:
    workbook = _build_workbook(report_data)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
